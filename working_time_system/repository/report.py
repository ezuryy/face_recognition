from openpyxl import load_workbook, worksheet
from typing import Tuple, Union, List

from fastapi import Depends
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta
import shutil
from uuid import UUID

from working_time_system.db.models import WorkTime, Request, Employee
from working_time_system.db.connection import get_session
from working_time_system.utils import RequestStatus

PATH_TO_REPORT = 'report'
PATH_TO_TEMPLATE = 'report/tabel13_example.xlsx'

EVENTS_CODES = {
    'weekend': 'В',
    'turnout': 'Я',
    'vacation': 'ОТ',
    'business_trip': 'К',
    'sick': 'Б',
    'strike': 'ЗБ'
}


class ReportRepository:
    def __init__(self, session: AsyncSession):
        self._session = session

        self.START_CELL_ROW = 24
        self.START_CELL_COL = 49

        self.CELL_WIDTH = 4
        self.CELL_HEIGHT = 1

        self.START_NAME_ROW = 24
        self.START_NAME_COL = 9

        self.NAME_WIDTH = 27
        self.NAME_HEIGHT = 4

    async def get_employee_id_by_name(self, name: str) -> UUID:
        query = select(Employee.id).where(Employee.name == name)
        employee_id = await self._session.scalar(query)
        return employee_id

    async def get_work_hours(self, request_date: datetime, name: str) -> int:
        query = select(WorkTime.work_hours).where((WorkTime.name == name) &
                                                  (func.DATE(WorkTime.date) == request_date.date())
                                                  )
        entry = await self._session.scalar(query)
        if entry is not None:
            return round(entry)
        return 0

    async def get_event(self, request_date: datetime, department: str, employee_id: UUID) -> Union[str, None]:
        # query = await self._session.execute(
        #     select(Request, Employee).join(
        #         Employee,
        #         Employee.id == Request.employee_id
        #     ).where(
        #         (Employee.department == department) &
        #         (Request.employee_id == employee_id) &
        #         (func.DATE(Request.from_dt) <= request_date.date()) &
        #         (func.DATE(Request.to_dt) >= request_date.date()) &
        #         (Request.status == Status.ok.name)
        #     )
        # )
        # event = query.scalars().first()
        query = select(Request.event).where(
            (Request.employee_id == employee_id) &
            (func.DATE(Request.from_dt) <= request_date.date()) &
            (func.DATE(Request.to_dt) >= request_date.date()) &
            (Request.status == RequestStatus.ok.name)
        )

        event = await self._session.scalar(query)
        if event is not None:
            return event.name
        return None

    async def set_name_value(
            self,
            sheet: worksheet,
            name: str,
            row: int,
    ) -> None:
        col = self.START_NAME_COL
        # col = self.START_CELL_COL + self.CELL_WIDTH * self.START_NAME_COL
        cell = sheet.cell(row=row, column=col)
        cell.value = name

    async def set_cell_value(
            self,
            sheet: worksheet,
            row_name_number: int,
            dt: datetime,
            event_type: str,
            work_hours: Union[int, None]
    ) -> None:
        if dt.date().day <= 15:
            row = row_name_number
            col_number = dt.date().day - 1
        else:
            row = row_name_number + 2
            col_number = dt.date().day - 15 - 1

        col = self.START_CELL_COL + self.CELL_WIDTH * col_number
        cell = sheet.cell(row=row, column=col)
        cell.value = event_type
        if event_type == 'Я':
            cell = sheet.cell(row=row + 1, column=col)
            cell.value = work_hours

    # async def get_names(self, sheet: worksheet) -> Tuple[List[str], List[int]]:
    #     row = self.START_NAME_ROW
    #     cell = sheet.cell(row, column=self.START_NAME_COL)
    #     names = []
    #     rows = []
    #     while cell.value is not None:
    #         names.append(cell.value)
    #         rows.append(row)
    #         row += self.CELL_HEIGHT
    #         cell = sheet.cell(row, column=self.START_NAME_COL)
    #     return names, rows

    async def get_report_names(self, employee_id: UUID, department: str) -> List[str]:
        query = await self._session.execute(select(Employee.name).where(
            (Employee.chief_id == employee_id) & (Employee.department == department)
        ))
        names = query.scalars().all()
        if len(names) == 0:
            query = await self._session.execute(select(Employee.name).where(Employee.id == employee_id))
            name = query.scalars().first()
            return [name]
        return names

    async def get_report(self, employee_id: UUID, department: str, month: int, year: int) -> str:
        date_from = datetime(year=year, month=month, day=1)
        date_to = (date_from + timedelta(days=32)).replace(day=1)
        filename = PATH_TO_REPORT + f'/report_{year}_{month}_{department}.xlsx'
        shutil.copy(PATH_TO_TEMPLATE, filename)
        wb = load_workbook(filename)
        sheet = wb['Т-13']
        names = await self.get_report_names(employee_id, department)
        # names_, rows = await self.get_names(sheet)
        while date_from < date_to:
            name_row = self.START_NAME_ROW
            for i, name in enumerate(names):
                # print(rows[i], row)
                await self.set_name_value(sheet, name, name_row)
                if date_from.weekday() >= 5:
                    await self.set_cell_value(sheet, name_row, date_from, 'B', None)
                    name_row += self.NAME_HEIGHT
                    continue

                employee_id = await self.get_employee_id_by_name(name)
                # request to Request table
                event = await self.get_event(date_from, department, employee_id)
                if event is not None:
                    await self.set_cell_value(sheet, name_row, date_from, EVENTS_CODES[event], None)
                    name_row += self.NAME_HEIGHT
                    continue

                # request to work_time table
                work_hours = await self.get_work_hours(date_from, name)
                if work_hours > 0:
                    await self.set_cell_value(sheet, name_row, date_from, 'Я', work_hours)
                name_row += self.NAME_HEIGHT

            date_from += timedelta(days=1)

        wb.save(filename)
        wb.close()
        return filename


async def get_report_repository(session: AsyncSession = Depends(get_session)) -> ReportRepository:
    return ReportRepository(session=session)
